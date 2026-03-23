"""
Excel / CSV Parser
===================
Parses messy real-world Indian SMB Excel and CSV files into the same
clean DataFrame format as the Tally XML parser.

Indian data quirks handled:
  - Mixed Hindi/English column names
  - Inconsistent date formats: DD/MM/YY, DD-MM-YYYY, "3 Jan 24", "03-Jan-2026"
  - Merged cells in Excel (forward-filled)
  - Blank header rows (auto-detected and skipped)
  - Values with ₹ symbol and Indian comma formatting (1,24,300)
  - Negative values that are returns, not errors
  - Multiple sheets — all sheets scanned, best-fit used

Returns a clean DataFrame with columns:
  date, customer, product, amount (Decimal), source_row, sheet_name
"""

import logging
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Candidate column name patterns for each semantic field.
# Comparison is lowercased + stripped. Order matters — first match wins.
COLUMN_PATTERNS: dict[str, list[str]] = {
    "date": [
        "date", "dt", "invoice date", "bill date", "transaction date",
        "sale date", "voucher date", "posting date", "entry date",
        "inv date", "txn date", "trans date", "document date",
        "challan date", "receipt date", "payment date", "booking date",
        "order date", "tarikh", "तारीख", "दिनांक",
    ],
    "customer": [
        "customer", "client", "party", "party name", "customer name",
        "buyer", "purchaser", "ledger name", "account name",
        "vendor", "supplier", "counterparty", "recipient",
        "consignee", "payer", "payee", "naam", "नाम", "ग्राहक", "khata",
    ],
    "product": [
        "product", "item", "description", "particulars", "goods",
        "article", "product name", "item name", "item description",
        "narration", "nature of supply", "service", "category",
        "account head", "sub ledger", "sku", "hsn code",
        "maal", "saman", "माल", "सामान", "वस्तु",
    ],
    "amount": [
        "amount", "total", "net amount", "sale amount", "revenue",
        "value", "net total", "invoice amount", "bill amount",
        "taxable value", "invoice value", "gross amount", "total amount",
        "transaction amount", "net value", "subtotal",
        "rashi", "राशि", "मूल्य", "कुल",
    ],
    "debit": [
        "debit", "dr amount", "debit amount",
        "withdrawal", "withdrawl", "outflow",
    ],
    "credit": [
        "credit", "cr amount", "credit amount",
        "deposit", "inflow", "received",
    ],
    "quantity": [
        "qty", "quantity", "units", "nos", "number", "count",
        "matra", "मात्रा", "pieces", "pcs",
    ],
    "unit_price": [
        "rate", "price", "unit price", "per unit", "mrp",
        "dar", "दर", "unit rate",
    ],
}

# Date format patterns to try in order (most common Indian formats first)
DATE_FORMATS: list[str] = [
    # Indian / day-first formats
    "%d/%m/%Y",       # 15/01/2026
    "%d-%m-%Y",       # 15-01-2026
    "%d/%m/%y",       # 15/01/26
    "%d-%m-%y",       # 15-01-26
    "%d/%m/%Y %H:%M", # 15/01/2026 08:26
    "%d/%m/%Y %H:%M:%S", # 15/01/2026 08:26:00
    # ISO
    "%Y-%m-%d",       # 2026-01-15
    "%Y-%m-%d %H:%M", # 2026-01-15 08:26
    "%Y-%m-%d %H:%M:%S", # 2026-01-15 08:26:00
    "%Y%m%d",         # 20260115 (Tally)
    # Month-name formats
    "%d %b %Y",       # 15 Jan 2026
    "%d %b %y",       # 15 Jan 26
    "%d-%b-%Y",       # 15-Jan-2026
    "%d-%b-%y",       # 15-Jan-26
    "%d/%b/%Y",       # 15/Jan/2026
    "%B %d, %Y",      # January 15, 2026
    # Dot-separated
    "%d.%m.%Y",       # 15.01.2026
    "%d.%m.%y",       # 15.01.26
    # American month-first formats (M/D/YYYY, MM/DD/YYYY with optional time)
    "%m/%d/%Y %H:%M", # 12/1/2010 8:26  — UCI Online Retail, US exports
    "%m/%d/%Y %H:%M:%S", # 12/1/2010 8:26:00
    "%m/%d/%Y",       # 12/01/2026
    "%m/%d/%y",       # 12/01/26
    "%m-%d-%Y",       # 12-01-2026
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class ExcelParseError(Exception):
    """Raised when the file cannot be parsed as a valid Excel/CSV export."""


class ExcelParseResult:
    """Container for parser output + metadata."""

    def __init__(
        self,
        df: pd.DataFrame,
        detected_columns: dict[str, str],
        sheets_processed: list[str],
        rows_dropped: int,
        warnings: list[str],
    ):
        self.df = df
        self.detected_columns = detected_columns   # semantic → actual column name
        self.sheets_processed = sheets_processed
        self.rows_dropped = rows_dropped
        self.warnings = warnings

    def __repr__(self) -> str:
        return (
            f"ExcelParseResult("
            f"rows={len(self.df)}, "
            f"sheets={self.sheets_processed}, "
            f"detected_columns={self.detected_columns}"
            f")"
        )


def parse_excel(
    source: str | Path | bytes | BytesIO,
    *,
    sheet_name: Optional[str] = None,
    column_map_override: Optional[dict] = None,
) -> ExcelParseResult:
    """
    Parse an Excel (.xlsx/.xls) file into a clean ExcelParseResult.

    Args:
        source:               File path, raw bytes, or BytesIO buffer.
        sheet_name:           Specific sheet to parse. None = auto-select best sheet.
        column_map_override:  Pre-detected column map (e.g. from Gemini) to use instead
                              of rule-based detection.

    Returns:
        ExcelParseResult with clean DataFrame and metadata.

    Raises:
        ExcelParseError: If no usable sales data can be found.
    """
    raw = _load_source(source)
    return _parse_workbook(raw, sheet_name=sheet_name, column_map_override=column_map_override)


def parse_csv(
    source: str | Path | bytes | BytesIO,
    *,
    column_map_override: Optional[dict] = None,
) -> ExcelParseResult:
    """
    Parse a CSV file into a clean ExcelParseResult.

    Handles common Indian CSV encodings: UTF-8, UTF-8 BOM, Latin-1.

    Args:
        source:               File path, raw bytes, or BytesIO buffer.
        column_map_override:  Pre-detected column map (e.g. from Gemini) to use instead
                              of rule-based detection.

    Returns:
        ExcelParseResult with clean DataFrame and metadata.
    """
    raw = _load_source(source)
    return _parse_csv_bytes(raw, column_map_override=column_map_override)


# ---------------------------------------------------------------------------
# Internal — Excel
# ---------------------------------------------------------------------------

def _load_source(source: str | Path | bytes | BytesIO) -> bytes:
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        return path.read_bytes()
    if isinstance(source, BytesIO):
        source.seek(0)
        return source.read()
    if isinstance(source, bytes):
        return source
    raise TypeError(f"Unsupported source type: {type(source)}")


def _parse_workbook(
    raw: bytes,
    sheet_name: Optional[str],
    column_map_override: Optional[dict] = None,
) -> ExcelParseResult:
    """Load workbook, pick best sheet, parse it."""
    buffer = BytesIO(raw)

    # Determine file type from magic bytes
    is_xlsx = raw[:4] == b"PK\x03\x04"  # ZIP signature = .xlsx
    engine = "openpyxl" if is_xlsx else "xlrd"

    try:
        xls = pd.ExcelFile(buffer, engine=engine)
    except Exception as exc:
        raise ExcelParseError(f"Cannot open Excel file: {exc}") from exc

    available_sheets = xls.sheet_names

    if sheet_name:
        if sheet_name not in available_sheets:
            raise ExcelParseError(
                f"Sheet {sheet_name!r} not found. Available: {available_sheets}"
            )
        sheets_to_try = [sheet_name]
    else:
        sheets_to_try = available_sheets

    best_result: Optional[tuple[pd.DataFrame, dict, str, list[str]]] = None
    all_warnings: list[str] = []

    for sname in sheets_to_try:
        try:
            raw_df = xls.parse(
                sname,
                header=None,    # We detect headers ourselves
                dtype=str,      # Read everything as string to avoid pandas type coercion
            )
        except Exception as exc:
            all_warnings.append(f"Could not read sheet {sname!r}: {exc}")
            continue

        # Forward-fill merged cells (openpyxl handles this for .xlsx)
        if is_xlsx:
            raw_df = _unmerge_cells(raw, sname, raw_df)

        header_row_idx = _find_header_row(raw_df)
        if header_row_idx is None:
            all_warnings.append(f"Sheet {sname!r}: no recognisable header row found — skipped")
            continue

        # Re-read with detected header
        df = xls.parse(sname, header=header_row_idx, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        column_map = _detect_columns(df.columns.tolist())

        # Override with Gemini-detected mapping if provided
        if column_map_override:
            # Only use columns that actually exist in the dataframe
            valid_override = {k: v for k, v in column_map_override.items() if v in df.columns}
            # Also carry _compute_expr (not a column name, no validation needed)
            if "_compute_expr" in column_map_override:
                valid_override["_compute_expr"] = column_map_override["_compute_expr"]
            if valid_override.get("date") or valid_override.get("amount") or valid_override.get("credit") or valid_override.get("_compute_expr"):
                # Merge: rule-based fills gaps Gemini doesn't cover (e.g. unit_price)
                column_map = {**column_map, **valid_override}

        has_amount = "amount" in column_map
        has_dr_cr = "debit" in column_map or "credit" in column_map
        has_qty_price = "quantity" in column_map and "unit_price" in column_map
        has_expr = "_compute_expr" in column_map
        if "date" not in column_map or (not has_amount and not has_dr_cr and not has_qty_price and not has_expr):
            all_warnings.append(
                f"Sheet {sname!r}: could not identify required date/amount columns — skipped. "
                f"Detected: {column_map}"
            )
            continue

        # Prefer sheet with most data
        if best_result is None or len(df) > len(best_result[0]):
            best_result = (df, column_map, sname, [])

    if best_result is None:
        raise ExcelParseError(
            "No usable sales data found in any sheet. "
            f"Sheets checked: {sheets_to_try}. Warnings: {all_warnings}"
        )

    df, column_map, chosen_sheet, sheet_warnings = best_result
    all_warnings.extend(sheet_warnings)

    clean_df, rows_dropped, parse_warnings = _normalise_dataframe(df, column_map, chosen_sheet)
    all_warnings.extend(parse_warnings)

    return ExcelParseResult(
        df=clean_df,
        detected_columns=column_map,
        sheets_processed=[chosen_sheet],
        rows_dropped=rows_dropped,
        warnings=all_warnings,
    )


def _unmerge_cells(raw: bytes, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """
    Forward-fill values across merged cells in .xlsx files.
    openpyxl reads merged cells as NaN after the first cell — we fill them.
    """
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        filled_df = pd.DataFrame(data)
        # Forward fill across columns (merged horizontal cells)
        filled_df = filled_df.ffill(axis=1)
        return filled_df
    except Exception:
        # Non-critical — fall back to original df
        return df


def _find_header_row(df: pd.DataFrame, max_scan_rows: int = 15) -> Optional[int]:
    """
    Scan the first N rows to find which row contains the column headers.

    A header row is identified by having the most cells that match our
    known column name patterns. Handles blank rows at top of file.
    """
    all_patterns = [p for patterns in COLUMN_PATTERNS.values() for p in patterns]

    best_row = None
    best_score = 0

    for i, row in df.head(max_scan_rows).iterrows():
        score = 0
        for cell in row:
            if pd.isna(cell):
                continue
            cell_lower = str(cell).strip().lower()
            if any(cell_lower == p or cell_lower in p or p in cell_lower for p in all_patterns):
                score += 1
        if score > best_score:
            best_score = score
            best_row = i

    return best_row if best_score >= 1 else None


def _detect_columns(columns: list[str]) -> dict[str, str]:
    """
    Map semantic fields to actual column names using pattern matching.

    Returns dict like: {"date": "Invoice Date", "amount": "Net Total", ...}
    """
    col_lower_map = {c.lower().strip(): c for c in columns}
    result: dict[str, str] = {}

    for semantic, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            # Exact match first — fastest and most reliable
            if pattern in col_lower_map:
                result[semantic] = col_lower_map[pattern]
                break
            # Partial match: pattern keyword appears inside the column name
            # (e.g., pattern "invoice date" found inside column "sales invoice date")
            # We do NOT do the reverse (column inside pattern) to avoid false positives
            # like "amount" matching pattern "dr amount".
            for col_lower, col_original in col_lower_map.items():
                if pattern in col_lower:
                    result[semantic] = col_original
                    break
            if semantic in result:
                break

    # Post-processing: if the "amount" column name contains "debit" or "credit",
    # it was misclassified via partial substring match. Remove it from "amount"
    # so that the debit/credit combination logic takes over.
    if "amount" in result:
        amt_lower = result["amount"].lower()
        if "debit" in amt_lower or "credit" in amt_lower:
            # Don't use this as the generic amount — the debit/credit pair handles it
            del result["amount"]
            # Ensure the column is also registered under the correct semantic
            if "debit" in amt_lower and "debit" not in result:
                result["debit"] = result["amount"] if "amount" in result else list(col_lower_map.values())[0]
            if "credit" in amt_lower and "credit" not in result:
                result["credit"] = result.get("amount", list(col_lower_map.values())[0])

    return result


def _eval_amount_expr(row: pd.Series, expr: str, df_columns: list[str]) -> Decimal:
    """
    Safely evaluate an arithmetic expression like 'UnitPrice * Quantity'.
    Only allows column names (from df_columns) and basic operators (+, -, *, /).
    No eval() — manually tokenizes and computes with Decimal arithmetic.
    """
    import re
    # Tokenize: extract column names and operators
    # Strategy: split by operators, keeping operators as tokens
    tokens = re.split(r"(\s*[\+\-\*\/]\s*)", expr)
    tokens = [t.strip() for t in tokens if t.strip()]

    values: list[Decimal] = []
    operators: list[str] = []

    for token in tokens:
        if token in ("+", "-", "*", "/"):
            operators.append(token)
        else:
            # Must be a column name
            if token in df_columns:
                raw = str(row.get(token, "")).strip()
                values.append(_parse_amount(raw))
            else:
                # Unknown token — try treating as a literal number
                try:
                    values.append(Decimal(token))
                except Exception:
                    logger.debug("Unknown token in expr %r: %r — using 0", expr, token)
                    values.append(Decimal(0))

    if not values:
        return Decimal(0)

    result = values[0]
    for i, op in enumerate(operators):
        if i + 1 >= len(values):
            break
        rhs = values[i + 1]
        if op == "*":
            result = result * rhs
        elif op == "/":
            result = result / rhs if rhs != Decimal(0) else Decimal(0)
        elif op == "+":
            result = result + rhs
        elif op == "-":
            result = result - rhs

    return result


def _validate_column_map(
    df: pd.DataFrame,
    column_map: dict[str, str],
    sample_size: int = 20,
) -> tuple[bool, str]:
    """
    Validate a column map by attempting to parse a sample of rows.
    Returns (is_valid, error_reason).
    A map is valid if >= 50% of sampled rows produce a parseable date AND non-zero amount.
    """
    sample = df.head(sample_size)
    date_col = column_map.get("date")
    amount_col = column_map.get("amount")
    compute_expr = column_map.get("_compute_expr")
    debit_col = column_map.get("debit")
    credit_col = column_map.get("credit")
    qty_col = column_map.get("quantity")
    unit_price_col = column_map.get("unit_price")
    df_columns = df.columns.tolist()

    use_dr_cr = not amount_col and (debit_col or credit_col)
    use_qty_price = not amount_col and not use_dr_cr and not compute_expr and bool(qty_col and unit_price_col)
    use_expr = bool(compute_expr) and not amount_col

    valid = 0
    total = 0
    date_failures = 0
    amount_zeros = 0

    for _, row in sample.iterrows():
        total += 1

        # Check date
        date_raw = str(row.get(date_col, "")).strip() if date_col else ""
        date = _parse_date(date_raw)
        if date is None:
            date_failures += 1
            continue

        # Check amount
        if use_expr and compute_expr:
            amount = _eval_amount_expr(row, compute_expr, df_columns)
        elif use_dr_cr:
            cr = _parse_amount(str(row.get(credit_col, "")).strip() if credit_col else "")
            dr = _parse_amount(str(row.get(debit_col, "")).strip() if debit_col else "")
            amount = cr - dr
        elif use_qty_price:
            qty = _parse_amount(str(row.get(qty_col, "")).strip() if qty_col else "")
            price = _parse_amount(str(row.get(unit_price_col, "")).strip() if unit_price_col else "")
            amount = qty * price
        else:
            amount_raw = str(row.get(amount_col, "")).strip() if amount_col else ""
            amount = _parse_amount(amount_raw)

        if amount == Decimal(0):
            amount_zeros += 1
            continue

        valid += 1

    if total == 0:
        return False, "No rows to validate"

    pct = valid / total
    if pct >= 0.5:
        return True, ""

    reasons = []
    if date_failures > total * 0.3:
        reasons.append(f"{date_failures}/{total} rows had unparseable dates in column {date_col!r}")
    if amount_zeros > total * 0.3:
        reasons.append(
            f"{amount_zeros}/{total} rows had zero amounts — "
            f"amount_col={amount_col!r}, compute_expr={compute_expr!r}"
        )
    if not reasons:
        reasons.append(f"only {valid}/{total} rows produced valid date+amount pairs")

    return False, "; ".join(reasons)


def _normalise_dataframe(
    df: pd.DataFrame,
    column_map: dict[str, str],
    sheet_name: str,
) -> tuple[pd.DataFrame, int, list[str]]:
    """
    Extract and normalise the mapped columns into our standard format.

    Returns: (clean_df, rows_dropped, warnings)
    """
    warnings: list[str] = []
    initial_rows = len(df)

    # Build standard columns
    clean: dict[str, list] = {
        "date": [],
        "customer": [],
        "product": [],
        "amount": [],
        "source_row": [],
        "sheet_name": [],
    }

    date_col = column_map.get("date")
    amount_col = column_map.get("amount")
    customer_col = column_map.get("customer")
    product_col = column_map.get("product")
    debit_col = column_map.get("debit")
    credit_col = column_map.get("credit")
    qty_col = column_map.get("quantity")
    unit_price_col = column_map.get("unit_price")
    compute_expr = column_map.get("_compute_expr")  # LLM-provided expression
    df_columns = df.columns.tolist()

    # Determine amount computation strategy (priority order)
    use_expr = bool(compute_expr) and not amount_col
    use_dr_cr = not amount_col and not use_expr and (debit_col or credit_col)
    use_qty_price = not amount_col and not use_expr and not use_dr_cr and bool(qty_col and unit_price_col)

    failed_dates = 0

    for idx, row in df.iterrows():
        # Parse date
        date_raw = str(row.get(date_col, "")).strip() if date_col else ""
        date = _parse_date(date_raw)
        if date is None:
            failed_dates += 1
            continue

        # Parse amount — LLM expression → debit/credit → qty×price → direct column
        if use_expr and compute_expr:
            amount = _eval_amount_expr(row, compute_expr, df_columns)
        elif use_dr_cr:
            cr = _parse_amount(str(row.get(credit_col, "")).strip() if credit_col else "")
            dr = _parse_amount(str(row.get(debit_col, "")).strip() if debit_col else "")
            # Net credit = income (positive), net debit = expense (negative)
            amount = cr - dr
        elif use_qty_price:
            qty = _parse_amount(str(row.get(qty_col, "")).strip() if qty_col else "")
            price = _parse_amount(str(row.get(unit_price_col, "")).strip() if unit_price_col else "")
            amount = qty * price
        else:
            amount_raw = str(row.get(amount_col, "")).strip() if amount_col else ""
            amount = _parse_amount(amount_raw)

        # Customer — optional, default to Walk-in
        customer_raw = str(row.get(customer_col, "")).strip() if customer_col else ""
        customer = customer_raw if customer_raw and customer_raw.lower() != "nan" else "Walk-in Customer"

        # Product — optional, default to "General"
        product_raw = str(row.get(product_col, "")).strip() if product_col else ""
        product = product_raw if product_raw and product_raw.lower() != "nan" else "General"

        clean["date"].append(date)
        clean["customer"].append(customer)
        clean["product"].append(product)
        clean["amount"].append(amount)
        clean["source_row"].append(int(idx) + 1)  # 1-indexed for user display
        clean["sheet_name"].append(sheet_name)

    if failed_dates > 0:
        warnings.append(
            f"Skipped {failed_dates} rows with unrecognisable dates "
            f"(out of {initial_rows} total rows)"
        )

    result_df = pd.DataFrame(clean)
    result_df["date"] = pd.to_datetime(result_df["date"])
    result_df["customer"] = result_df["customer"].astype("string")
    result_df["product"] = result_df["product"].astype("string")
    result_df["sheet_name"] = result_df["sheet_name"].astype("string")

    # Filter out zero-amount rows (header duplicates, subtotals)
    zero_mask = result_df["amount"].apply(lambda x: x == Decimal(0))
    result_df = result_df[~zero_mask].reset_index(drop=True)

    rows_dropped = initial_rows - len(result_df)

    if result_df.empty or "date" not in result_df.columns:
        raise ExcelParseError(
            f"No valid rows could be parsed. "
            f"All {initial_rows} rows failed date/amount validation. "
            f"Check that the file has recognisable date and amount columns."
        )

    result_df = result_df.sort_values("date").reset_index(drop=True)

    return result_df, rows_dropped, warnings


# ---------------------------------------------------------------------------
# Internal — CSV
# ---------------------------------------------------------------------------

def _parse_csv_bytes(raw: bytes, column_map_override: Optional[dict] = None) -> ExcelParseResult:
    """Parse CSV bytes with automatic encoding detection."""
    warnings: list[str] = []

    # Try encodings in order of likelihood for Indian data
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ExcelParseError("Could not decode CSV — unknown encoding")

    buffer = BytesIO(text.encode("utf-8"))

    # Try different separators
    for sep in (",", "\t", ";", "|"):
        try:
            df = pd.read_csv(buffer, sep=sep, dtype=str, header=None)
            if df.shape[1] >= 2:  # At least 2 columns
                buffer.seek(0)
                break
        except Exception:
            buffer.seek(0)
            continue
    else:
        raise ExcelParseError("Could not parse CSV — unrecognised delimiter")

    buffer.seek(0)
    df_full = pd.read_csv(buffer, sep=sep, dtype=str, header=None)

    header_row_idx = _find_header_row(df_full)
    if header_row_idx is None:
        raise ExcelParseError("No header row found in CSV file")

    buffer.seek(0)
    df = pd.read_csv(buffer, sep=sep, dtype=str, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]

    column_map = _detect_columns(df.columns.tolist())

    # Override with Gemini-detected mapping if provided
    if column_map_override:
        valid_override = {k: v for k, v in column_map_override.items() if v in df.columns}
        # Also carry _compute_expr (not a column name, no validation needed)
        if "_compute_expr" in column_map_override:
            valid_override["_compute_expr"] = column_map_override["_compute_expr"]
        if valid_override.get("date") or valid_override.get("amount") or valid_override.get("credit") or valid_override.get("_compute_expr"):
            # Merge: rule-based fills gaps Gemini doesn't cover (e.g. unit_price)
            column_map = {**column_map, **valid_override}

    has_amount = "amount" in column_map
    has_dr_cr = "debit" in column_map or "credit" in column_map
    has_qty_price = "quantity" in column_map and "unit_price" in column_map
    has_expr = "_compute_expr" in column_map
    if "date" not in column_map or (not has_amount and not has_dr_cr and not has_qty_price and not has_expr):
        raise ExcelParseError(
            f"Could not detect required date/amount columns. "
            f"Detected: {column_map}. Columns found: {df.columns.tolist()}"
        )

    clean_df, rows_dropped, parse_warnings = _normalise_dataframe(df, column_map, "CSV")
    warnings.extend(parse_warnings)

    return ExcelParseResult(
        df=clean_df,
        detected_columns=column_map,
        sheets_processed=["CSV"],
        rows_dropped=rows_dropped,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _parse_date(raw: str) -> Optional[pd.Timestamp]:
    """
    Parse an Indian date string into a pandas Timestamp.

    Tries formats in order from most to least common.
    Returns None if no format matches (caller decides whether to skip row).
    """
    if not raw or raw.lower() in ("nan", "none", "null", "-", ""):
        return None

    # Clean common noise
    raw = raw.strip()
    raw = re.sub(r"\s+", " ", raw)

    for fmt in DATE_FORMATS:
        try:
            return pd.Timestamp(pd.to_datetime(raw, format=fmt))
        except (ValueError, TypeError):
            continue

    # Last resort: let pandas infer using format='mixed' (pandas 2.x compatible)
    try:
        return pd.Timestamp(pd.to_datetime(raw, format="mixed", dayfirst=True))
    except Exception:
        pass

    # Final fallback: plain parse with no hints
    try:
        return pd.Timestamp(pd.to_datetime(raw))
    except Exception:
        logger.debug("Could not parse date: %r", raw)
        return None


def _parse_amount(raw: str) -> Decimal:
    """
    Parse an Indian-formatted amount string into Python Decimal.

    Handles:
    - ₹ symbol: "₹1,24,300" → Decimal("124300")
    - Indian comma grouping: "1,24,300" → Decimal("124300")
    - Parentheses for negatives: "(5000)" → Decimal("-5000")
    - CR/DR suffixes: "5000 DR" → Decimal("5000"), "5000 CR" → Decimal("-5000")
    - Plain numbers: "15000.50" → Decimal("15000.50")
    """
    if not raw or raw.lower() in ("nan", "none", "null", "-", ""):
        return Decimal(0)

    raw = raw.strip()

    # Parentheses = negative (accounting convention)
    is_negative = raw.startswith("(") and raw.endswith(")")
    if is_negative:
        raw = raw[1:-1]

    # Handle CR/DR suffix
    if raw.upper().endswith(" CR"):
        raw = raw[:-3]
        is_negative = True
    elif raw.upper().endswith(" DR"):
        raw = raw[:-3]

    # Remove currency symbols and commas
    cleaned = re.sub(r"[₹$,\s]", "", raw)

    if not cleaned:
        return Decimal(0)

    try:
        value = Decimal(cleaned)
        return -value if is_negative else value
    except InvalidOperation:
        logger.debug("Could not parse amount: %r", raw)
        return Decimal(0)


# ---------------------------------------------------------------------------
# Peek helpers — for Gemini schema detection pre-pass
# ---------------------------------------------------------------------------

def peek_raw_sample(
    raw: bytes,
    max_rows: int = 8,
) -> tuple[list[str], list[dict]]:
    """
    Quickly read headers + sample rows without normalization.
    Used by Gemini schema detector before full parsing.
    Returns (headers, sample_rows_as_dicts).
    """
    is_xlsx = raw[:4] == b"PK\x03\x04"
    engine = "openpyxl" if is_xlsx else "xlrd"
    buffer = BytesIO(raw)
    try:
        xls = pd.ExcelFile(buffer, engine=engine)
        sname = xls.sheet_names[0]
        raw_df = xls.parse(sname, header=None, dtype=str)
        header_idx = _find_header_row(raw_df)
        if header_idx is None:
            return [], []
        df = xls.parse(sname, header=header_idx, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        sample = df.head(max_rows).fillna("").to_dict(orient="records")
        return df.columns.tolist(), sample
    except Exception:
        return [], []


def peek_raw_sample_csv(
    raw: bytes,
    max_rows: int = 8,
) -> tuple[list[str], list[dict]]:
    """Same as peek_raw_sample but for CSV bytes."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return [], []

    for sep in (",", "\t", ";", "|"):
        try:
            df_raw = pd.read_csv(BytesIO(text.encode("utf-8")), sep=sep, dtype=str, header=None)
            if df_raw.shape[1] >= 2:
                break
        except Exception:
            continue
    else:
        return [], []

    df_raw = pd.read_csv(BytesIO(text.encode("utf-8")), sep=sep, dtype=str, header=None)
    header_idx = _find_header_row(df_raw)
    if header_idx is None:
        return [], []

    df = pd.read_csv(BytesIO(text.encode("utf-8")), sep=sep, dtype=str, header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    sample = df.head(max_rows).fillna("").to_dict(orient="records")
    return df.columns.tolist(), sample
